#--------------------------------------------------------------------------------------------------------------------
#   Define a Python library.
#--------------------------------------------------------------------------------------------------------------------
import os
import xlrd #for read excel workbook [READ ONLY!]
from openpyxl.styles import PatternFill, Alignment
from openpyxl import Workbook, load_workbook
import operator
import collections

#--------------------------------------------------------------------------------------------------------------------
#   Define a Functions
#--------------------------------------------------------------------------------------------------------------------
def xlsxCreate_TN_3AS_PC(savePath):

    workbook = Workbook()
    sheeto = workbook.active

    sheeto["A1"] = "Tracking number"
    sheeto["B1"] = "เบอร์โทรศัพท์"
    sheeto["C1"] = "แขวง/ตำบล"
    sheeto["D1"] = "เขต/อำเภอ"
    sheeto["E1"] = "จังหวัด"
    sheeto["F1"] = "รหัสไปรษณีย์"

    sheeto.row_dimensions[1].height = 20
    sheeto.column_dimensions['A'].width = 18
    sheeto.column_dimensions['B'].width = 15
    sheeto.column_dimensions['C'].width = 24
    sheeto.column_dimensions['D'].width = 24
    sheeto.column_dimensions['E'].width = 24
    sheeto.column_dimensions['F'].width = 18
    
    blue_fill = PatternFill(start_color='99FFFF', end_color='99FFFF', fill_type='solid')
    for step in range(6):
        sheeto[Number_of_cell_alphabet(step + 1) + str(1)].fill = blue_fill
        sheeto[Number_of_cell_alphabet(step + 1) + str(1)].alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(filename = savePath)
    
    return()
def Number_of_cell_alphabet(alphabetNumber):

    if alphabetNumber == 1:
        thisAlphabet = 'A'
    elif alphabetNumber == 2:
        thisAlphabet = 'B'
    elif alphabetNumber == 3:
        thisAlphabet = 'C'
    elif alphabetNumber == 4:
        thisAlphabet = 'D'
    elif alphabetNumber == 5:
        thisAlphabet = 'E'
    elif alphabetNumber == 6:
        thisAlphabet = 'F'
    elif alphabetNumber == 7:
        thisAlphabet = 'G'
    elif alphabetNumber == 8:
        thisAlphabet = 'H'
    elif alphabetNumber == 9:
        thisAlphabet = 'I'
    elif alphabetNumber == 10:
        thisAlphabet = 'J'
    elif alphabetNumber == 11:
        thisAlphabet = 'K'
    elif alphabetNumber == 12:
        thisAlphabet = 'L'
    elif alphabetNumber == 13:
        thisAlphabet = 'M'
    elif alphabetNumber == 14:
        thisAlphabet = 'N'
    elif alphabetNumber == 15:
        thisAlphabet = 'O'
    elif alphabetNumber == 16:
        thisAlphabet = 'P'
    elif alphabetNumber == 17:
        thisAlphabet = 'Q'
    elif alphabetNumber == 18:
        thisAlphabet = 'R'
    elif alphabetNumber == 19:
        thisAlphabet = 'S'
    elif alphabetNumber == 20:
        thisAlphabet = 'T'
    elif alphabetNumber == 21:
        thisAlphabet = 'U'
    elif alphabetNumber == 22:
        thisAlphabet = 'V'
    elif alphabetNumber == 23:
        thisAlphabet = 'W'
    elif alphabetNumber == 24:
        thisAlphabet = 'X'
    elif alphabetNumber == 25:
        thisAlphabet = 'Y'
    elif alphabetNumber == 26:
        thisAlphabet = 'Z'
    else:
        print("Uncorrect alphabet number !")
    
    if thisAlphabet:
        return(thisAlphabet)
def Find_name_amonut(path):

    for files in os.walk(path): #for root, dirs, files in os.walk("./Checking File"):
        for listOfFiles in files:
            #Nothing.
            pass

    return(listOfFiles, len(listOfFiles))

#--------------------------------------------------------------------------------------------------------------------
#   Define a constant variables.
#--------------------------------------------------------------------------------------------------------------------
rawFilesLocation = "Excel files/"
savePath = "แยกข้อมูล.xlsx"
recordingOrder = 2 #start recording at the second row.

#--------------------------------------------------------------------------------------------------------------------
#   Create & Open & Prepair files.
#--------------------------------------------------------------------------------------------------------------------
xlsxCreate_TN_3AS_PC(savePath)

recorderWorkbook = load_workbook(savePath)  #for [Openpyxl] library
sheeto = recorderWorkbook.active

listOfFiles, filesAmount = Find_name_amonut(rawFilesLocation)
#--------------------------------------------------------------------------------------------------------------------
#   Read -> Separate -> Record contents.
#--------------------------------------------------------------------------------------------------------------------
print("----------------------------------------------------------------------------------------------------------\n")
for files in range(filesAmount):

    #Dynamic opening.
    workbookVar = xlrd.open_workbook(rawFilesLocation + listOfFiles[files])   #for [xlrd] library
    readerVar = workbookVar.sheet_by_index(0)

    for recordingStep in range (len(readerVar.col_values(0)) - 2):
        workingRow = recordingStep + 1
        addressLength = len(str(readerVar.cell(workingRow, 5)))
        strTemp = ""
        indexPicker = addressLength - 11 #11 because of the file's format (cutted a "thailand" text.)
        temp = str(readerVar.cell(workingRow, 5))[indexPicker]
        for i in range(3):
            while(temp != ' '):
                strTemp = strTemp + temp
                indexPicker -= 1
                temp = str(readerVar.cell(workingRow, 5))[indexPicker]

            if(i==0):
                sheeto['A' + str(recordingOrder)] = str(readerVar.cell(workingRow, 2).value)
                sheeto['B' + str(recordingOrder)] = str(readerVar.cell(workingRow, 4).value)
                sheeto['E' + str(recordingOrder)] = strTemp[::-1]
            elif(i==1):
                sheeto['C' + str(recordingOrder)] = strTemp[::-1]
            else:
                sheeto['D' + str(recordingOrder)] = strTemp[::-1]

            strTemp = ""
            indexPicker -= 1
            temp = str(readerVar.cell(workingRow, 5))[indexPicker]

        recordingOrder += 1
    print(" completed (" + str(files + 1) + "/" + str(filesAmount) + ") : " + listOfFiles[files])
#print("\n----------------------------------------------------------------------------------------------------------")
#--------------------------------------------------------------------------------------------------------------------
recorderWorkbook.save(filename = savePath)
print("\n -> Finished.\n")
print("----------------------------------------------------------------------------------------------------------")

#--------------------------------------------------------------------------------------------------------------------
#   Data Analysis || Additional fuction.
#--------------------------------------------------------------------------------------------------------------------

#create 3 sheets
recorderWorkbook.create_sheet("อันดับ-แขวง")
recorderWorkbook.create_sheet("อันดับ-เขต")
recorderWorkbook.create_sheet("อันดับ-จังหวัด")

sheeto_SD = recorderWorkbook["อันดับ-แขวง"]
sheeto_D = recorderWorkbook["อันดับ-เขต"]
sheeto_P = recorderWorkbook["อันดับ-จังหวัด"]

sheeto_SD["A1"], sheeto_D["A1"], sheeto_P["A1"] = "อันดับ", "อันดับ", "อันดับ"
sheeto_SD["B1"] = "แขวง"
sheeto_D["B1"] = "เขต"
sheeto_P["B1"] = "จังหวัด"
sheeto_SD["C1"], sheeto_D["C1"], sheeto_P["C1"] = "จำนวนลูกค้า", "จำนวนลูกค้า", "จำนวนลูกค้า"

#sheet styling
rankingSheet = [sheeto_SD, sheeto_D, sheeto_P]

blue_fill = PatternFill(start_color='99FFFF', end_color='99FFFF', fill_type='solid')

for sheetTopic in rankingSheet:
    sheetTopic.row_dimensions[1].height = 20
    sheetTopic.column_dimensions['B'].width = 24
    sheetTopic.column_dimensions['C'].width = 18
    for step in range(3):
        sheetTopic[Number_of_cell_alphabet(step + 1) + str(1)].fill = blue_fill
        sheetTopic[Number_of_cell_alphabet(step + 1) + str(1)].alignment = Alignment(horizontal='center', vertical='center')

#list variables
subDistrict = []
district = []
province = []

#dictionary variables
subDistrict_dic = {}
district_dic = {}
province_dic = {}

#fill data to 3 lists
for order in range(recordingOrder):
    subDistrict.append(sheeto['C'+ str(order+2)].value)
    district.append(sheeto['D'+ str(order+2)].value)
    province.append(sheeto['E'+ str(order+2)].value)

#fill data to 3 dictionaries
for aSD in subDistrict:
  if aSD in subDistrict_dic:
    subDistrict_dic[aSD] += 1
  else:
    subDistrict_dic[aSD] = 1

for aD in district:
  if aD in district_dic:
    district_dic[aD] += 1
  else:
    district_dic[aD] = 1

for aP in province:
  if aP in province_dic:
    province_dic[aP] += 1
  else:
    province_dic[aP] = 1

#pre-processing before fill into the cell | sorting by customer amount.
subDistrict_dic_sorted = collections.OrderedDict(sorted(subDistrict_dic.items(), key=operator.itemgetter(1), reverse=True))
district_dic_sorted = collections.OrderedDict(sorted(district_dic.items(), key=operator.itemgetter(1), reverse=True))
province_dic_sorted = collections.OrderedDict(sorted(province_dic.items(), key=operator.itemgetter(1), reverse=True))


#Fill in excel-----------------------------------------------------------------
rankingSortedDic = [subDistrict_dic_sorted, district_dic_sorted, province_dic_sorted]

for rankingType in range(3):
    rank = 1
    for candidate in rankingSortedDic[rankingType]:
        rankingSheet[rankingType]['A'+str(rank+1)] = rank
        rankingSheet[rankingType]['A'+str(rank+1)].alignment = Alignment(horizontal='center')
        rankingSheet[rankingType]['B'+str(rank+1)] = candidate
        rankingSheet[rankingType]['C'+str(rank+1)] = rankingSortedDic[rankingType].get(candidate)
        rank += 1

recorderWorkbook.save(filename = savePath)
